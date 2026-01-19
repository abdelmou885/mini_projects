import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import csv
import sys


def sync_excel_files(database_path, export_path, sheet_name="charm", output_path=None):
    """
    Synchronize database Excel file with export CSV file based on Change ID.

    Args:
        database_path: Path to the database Excel file (the one to be updated)
        export_path: Path to the export CSV file (source of truth for what should exist)
        sheet_name: Name of the sheet in the database Excel file to update (default: "charm")
        output_path: Path to save the updated file (optional, defaults to database_path)
    """

    # Validate files exist
    import os
    if not os.path.exists(database_path):
        print(f"Error: Database file not found: {database_path}")
        return
    if not os.path.exists(export_path):
        print(f"Error: Export file not found: {export_path}")
        return

    print(f"Database file: {database_path}")
    print(f"Export file: {export_path}")
    print(f"Target sheet: {sheet_name}")

    # Load database workbook
    print("\nLoading Excel file...")
    try:
        print("Loading database file...")
        db_wb = openpyxl.load_workbook(database_path)
        print("✓ Database file loaded successfully")
    except Exception as e:
        print(f"✗ Error loading database file: {str(e)}")
        print("\nTroubleshooting:")
        print("1. Open the file in Excel and verify it opens correctly")
        print("2. Save it as a new .xlsx file (File > Save As > Excel Workbook)")
        print("3. Make sure the file is not open in Excel")
        return

    # Check if the target sheet exists
    if sheet_name not in db_wb.sheetnames:
        print(f"\n✗ Error: Sheet '{sheet_name}' not found in database file")
        print(f"Available sheets: {', '.join(db_wb.sheetnames)}")
        return

    db_ws = db_wb[sheet_name]
    print(f"✓ Found sheet '{sheet_name}'")

    # Load CSV export file
    print("\nLoading CSV export file...")
    try:
        export_data = []
        with open(export_path, 'r', encoding='utf-8-sig') as csvfile:
            csv_reader = csv.reader(csvfile)
            for row in csv_reader:
                export_data.append(row)
        print(f"✓ CSV file loaded successfully ({len(export_data)} rows)")
    except Exception as e:
        print(f"✗ Error loading CSV file: {str(e)}")
        print("\nTrying with different encoding...")
        try:
            with open(export_path, 'r', encoding='latin-1') as csvfile:
                csv_reader = csv.reader(csvfile)
                export_data = list(csv_reader)
            print(f"✓ CSV file loaded successfully ({len(export_data)} rows)")
        except Exception as e2:
            print(f"✗ Error loading CSV file: {str(e2)}")
            return

    if len(export_data) == 0:
        print("✗ Error: CSV file is empty")
        return

    # Find column indices in database sheet
    print("Analyzing column structure...")

    # Get database columns
    db_headers = {}
    for col in range(1, db_ws.max_column + 1):
        header = db_ws.cell(1, col).value
        if header:
            db_headers[header.strip().lower()] = col

    # Get export CSV columns (first row)
    export_headers = {}
    if export_data:
        for col_idx, header in enumerate(export_data[0], start=1):
            if header:
                export_headers[header.strip().lower()] = col_idx

    # Verify required columns exist
    required_cols = ['change id', 'priority', 'description']

    for col in required_cols:
        if col not in db_headers:
            print(f"Error: '{col}' column not found in database file")
            return
        if col not in export_headers:
            print(f"Error: '{col}' column not found in export file")
            return

    # Optional columns to sync (if they exist in export)
    optional_cols = ['status']
    cols_to_sync = required_cols + [col for col in optional_cols if col in export_headers]

    print(f"Columns to sync: {', '.join(cols_to_sync)}")

    # Build map of Change IDs in database (row number -> change_id)
    print("\nScanning database for existing Change IDs...")
    db_change_ids = {}  # {change_id: row_number}
    for row in range(2, db_ws.max_row + 1):
        change_id = db_ws.cell(row, db_headers['change id']).value
        if change_id:
            db_change_ids[str(change_id).strip()] = row

    print(f"Found {len(db_change_ids)} existing Change IDs in database")

    # Build map of Change IDs in CSV export file
    print("Scanning CSV export file for Change IDs...")
    export_change_ids = {}  # {change_id: {col_name: value}}
    for row_idx in range(1, len(export_data)):  # Skip header row
        row = export_data[row_idx]
        if len(row) >= export_headers['change id']:
            change_id = row[export_headers['change id'] - 1]  # -1 because list is 0-indexed
            if change_id:
                change_id_str = str(change_id).strip()
                export_change_ids[change_id_str] = {}
                for col_name in cols_to_sync:
                    if col_name in export_headers:
                        col_idx = export_headers[col_name] - 1  # -1 for 0-indexed list
                        if col_idx < len(row):
                            export_change_ids[change_id_str][col_name] = row[col_idx]

    print(f"Found {len(export_change_ids)} Change IDs in export file")

    # Scenario 3: Delete rows that exist in database but not in export
    print("\n=== Scenario 3: Removing resolved tickets ===")
    rows_to_delete = []
    for change_id, row_num in db_change_ids.items():
        if change_id not in export_change_ids:
            rows_to_delete.append(row_num)
            print(f"Will delete Change ID {change_id} (row {row_num})")

    # Delete rows in reverse order to avoid index shifting
    for row_num in sorted(rows_to_delete, reverse=True):
        db_ws.delete_rows(row_num, 1)
        print(f"Deleted row {row_num}")

    if rows_to_delete:
        print(f"Deleted {len(rows_to_delete)} resolved tickets")
    else:
        print("No tickets to delete")

    # Rebuild database change_ids map after deletion
    db_change_ids = {}
    for row in range(2, db_ws.max_row + 1):
        change_id = db_ws.cell(row, db_headers['change id']).value
        if change_id:
            db_change_ids[str(change_id).strip()] = row

    # Scenario 1: IDs exist in both - keep existing (do nothing)
    print("\n=== Scenario 1: Keeping existing tickets ===")
    existing_count = 0
    for change_id in export_change_ids:
        if change_id in db_change_ids:
            existing_count += 1
            print(f"Change ID {change_id} exists in both - keeping existing data")

    if existing_count:
        print(f"Kept {existing_count} existing tickets unchanged")

    # Scenario 2: IDs exist in export but not in database - insert at end
    print("\n=== Scenario 2: Adding new tickets ===")
    white_fill = PatternFill(fill_type=None)  # No fill (white)
    new_tickets = []

    for change_id, data in export_change_ids.items():
        if change_id not in db_change_ids:
            new_tickets.append((change_id, data))

    if new_tickets:
        next_row = db_ws.max_row + 1
        for change_id, data in new_tickets:
            print(f"Adding Change ID {change_id} to row {next_row}")

            # Insert data for synced columns
            for col_name in cols_to_sync:
                if col_name in db_headers and col_name in data:
                    col_idx = db_headers[col_name]
                    cell = db_ws.cell(next_row, col_idx)
                    cell.value = data[col_name]
                    # Set white fill (no background color)
                    cell.fill = white_fill

            next_row += 1

        print(f"Added {len(new_tickets)} new tickets")
    else:
        print("No new tickets to add")

    # Save the updated workbook
    output_file = output_path if output_path else database_path
    print(f"\nSaving updated database to: {output_file}")
    db_wb.save(output_file)
    print("✓ Synchronization complete!")

    # Summary
    print("\n" + "=" * 50)
    print("SUMMARY")
    print("=" * 50)
    print(f"Tickets kept unchanged: {existing_count}")
    print(f"Tickets added: {len(new_tickets)}")
    print(f"Tickets deleted: {len(rows_to_delete)}")
    print("=" * 50)


if __name__ == "__main__":
    import os

    # Get the directory where this script is located
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # ========================================
    # UPDATE THESE FILENAMES
    # ========================================

    DATABASE_FILENAME = "transport.xlsx"  # ← Your Excel database file
    EXPORT_FILENAME = "export.csv"  # ← Your CSV export file
    SHEET_NAME = "charm"  # ← Sheet name in Excel file
    OUTPUT_FILENAME = "transport_updated.xlsx"  # ← (Optional) Output filename

    # Build full paths (files in same folder as script)
    DATABASE_FILE = os.path.join(script_dir, DATABASE_FILENAME)
    EXPORT_FILE = os.path.join(script_dir, EXPORT_FILENAME)
    OUTPUT_FILE = os.path.join(script_dir, OUTPUT_FILENAME)

    print(f"Script location: {script_dir}")
    print(f"Looking for files in: {script_dir}\n")

    # ========================================
    # CHOOSE ONE OPTION BELOW AND UNCOMMENT IT
    # ========================================

    # OPTION 1: Update database file directly (overwrites original)
    sync_excel_files(DATABASE_FILE, EXPORT_FILE, SHEET_NAME)

    # OPTION 2: Save to a new file (keeps original safe)
    # sync_excel_files(DATABASE_FILE, EXPORT_FILE, SHEET_NAME, OUTPUT_FILE)